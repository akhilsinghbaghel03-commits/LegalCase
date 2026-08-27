"""
report_generator.py - Comprehensive, Professional Test Reporting Engine.
Generates:
1. Multi-sheet Excel Report (Execution Summary, Test Case Details, Module Summary, Failed Test Cases)
2. Standalone, Responsive, Modern HTML Report (Dashboard KPIs, Embedded SVG Charts, Real-time Search/Filter, Collapsible Traces)
3. Self-Contained, Email-Optimized HTML Body (Execution Summary, Metrics, Failed Cases Table, Jenkins Info)
4. summary.json for pipeline data extraction.
"""

import os
import sys
import re
import json
import html
import datetime
import xml.etree.ElementTree as ET

# Ensure openpyxl is available
try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    import subprocess
    print("openpyxl not found. Installing openpyxl...")
    subprocess.check_call([sys.executable, "-m", "pip", "install", "openpyxl"])
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter


class ReportGenerator:
    def __init__(self, source_path=None, output_dir="reports"):
        self.output_dir = output_dir
        os.makedirs(self.output_dir, exist_ok=True)

        # Dynamic metadata from environment
        self.project_name = os.environ.get("PROJECT_NAME", "YorPro")
        self.environment = os.environ.get("ENVIRONMENT", "DEV")
        self.build_number = os.environ.get("BUILD_NUMBER", "Local")
        self.build_url = os.environ.get("BUILD_URL", "")
        self.job_name = os.environ.get("JOB_NAME", "Legalhub")
        self.browser = os.environ.get("BROWSER", "Chrome")

        # Execution times
        self.now = datetime.datetime.now()
        self.execution_date = self.now.strftime("%Y-%m-%d")
        self.start_time_str = self.now.strftime("%Y-%m-%d %H:%M:%S")

        # Parsed test data
        self.test_cases = []
        self.modules = {}
        self.total_tests = 0
        self.passed_tests = 0
        self.failed_tests = 0
        self.skipped_tests = 0
        self.total_duration = 0.0
        self.overall_status = "PASSED"

        self.source_path = source_path
        self._load_and_parse()

    def _clean_module_name(self, classname_or_id):
        """Derive a friendly, professional module name."""
        raw = str(classname_or_id or "General")
        if "test_contact" in raw.lower():
            return "Contact Management Module"
        elif "test_matter" in raw.lower():
            return "Matter & Case Management"
        elif "test_registration" in raw.lower() or "signup" in raw.lower():
            return "User Registration & Onboarding"
        elif "test_login" in raw.lower() or "invalid_login" in raw.lower():
            return "Authentication & Security"
        elif "test_forgot" in raw.lower():
            return "Password Recovery & Reset"
        else:
            base = raw.split("::")[0].split("/")[-1].split("\\")[-1].replace("test_", "").replace(".py", "").replace("_", " ").title()
            return f"{base} Module" if not base.endswith("Module") else base

    def _clean_scenario_name(self, name_or_id):
        """Format test function name into a readable scenario."""
        raw = str(name_or_id)
        if "::" in raw:
            raw = raw.split("::")[-1]
        
        param = ""
        if "[" in raw and raw.endswith("]"):
            base, param = raw.split("[", 1)
            param = param[:-1]
            raw = base

        clean = raw.replace("test_", "").replace("_", " ").strip().capitalize()
        if param:
            return f"{clean} (Input: {param})"
        return clean

    def _parse_duration_str(self, dur_str):
        """Convert duration string to seconds float."""
        if not dur_str:
            return 0.0
        dur_str = str(dur_str).strip()
        try:
            if "ms" in dur_str:
                return round(float(dur_str.replace("ms", "").strip()) / 1000.0, 3)
            elif ":" in dur_str:
                parts = dur_str.split(":")
                if len(parts) == 3:
                    return round(float(parts[0]) * 3600 + float(parts[1]) * 60 + float(parts[2]), 2)
                elif len(parts) == 2:
                    return round(float(parts[0]) * 60 + float(parts[1]), 2)
            else:
                return round(float(dur_str.replace("s", "").strip()), 2)
        except Exception:
            return 0.0
        return 0.0

    def _load_and_parse(self):
        """Parse test results from report_full.html, reports/results.xml, or reports/report.html."""
        candidates = []
        if self.source_path and os.path.exists(self.source_path):
            candidates.append(self.source_path)
        candidates.extend([
            "report_full.html",
            os.path.join(self.output_dir, "results.xml"),
            "reports/results.xml",
            os.path.join(self.output_dir, "report.html"),
            "reports/report.html"
        ])

        parsed = False
        for c_path in candidates:
            if os.path.exists(c_path):
                if c_path.endswith(".html"):
                    parsed = self._parse_pytest_html(c_path)
                elif c_path.endswith(".xml"):
                    parsed = self._parse_junit_xml(c_path)
                if parsed:
                    print(f"[INFO] Successfully loaded test results from: {c_path}")
                    break

        if not parsed:
            print("[WARNING] No valid test result artifact found. Initialized with empty state.")

        self.overall_status = "PASSED" if (self.failed_tests == 0 and self.total_tests > 0) else ("FAILED" if self.failed_tests > 0 else "UNKNOWN")
        self.end_time_str = (self.now + datetime.timedelta(seconds=self.total_duration)).strftime("%Y-%m-%d %H:%M:%S")

    def _parse_pytest_html(self, html_path):
        """Extract test data from pytest-html data blob."""
        try:
            with open(html_path, "r", encoding="utf-8") as f:
                content = f.read()

            match = re.search(r'data-jsonblob="([^"]+)"', content)
            if not match:
                return False

            raw_json = html.unescape(match.group(1))
            data = json.loads(raw_json)
            tests_dict = data.get("tests", {})
            if not tests_dict:
                return False

            tc_index = 1
            for test_key, items in tests_dict.items():
                for item in items:
                    res_raw = item.get("result", "Passed").upper()
                    status = "PASS" if "PASS" in res_raw else ("FAIL" if "FAIL" in res_raw or "ERROR" in res_raw else "SKIPPED")
                    dur_str = item.get("duration", "0")
                    duration = self._parse_duration_str(dur_str)
                    self.total_duration += duration

                    log_text = item.get("log", "")
                    if log_text == "No log output captured.":
                        log_text = ""

                    error_reason = ""
                    if status == "FAIL":
                        error_reason = log_text.split("\n")[0] if log_text else "Test failed"
                    elif status == "SKIPPED":
                        error_reason = log_text.strip() or "Test marked skipped"

                    test_id = item.get("testId", test_key)
                    module_name = self._clean_module_name(test_key)
                    scenario_name = self._clean_scenario_name(test_key)

                    tc_data = {
                        "id": f"TC_{tc_index:03d}",
                        "name": test_key.split("::")[-1] if "::" in test_key else test_key,
                        "module": module_name,
                        "scenario": scenario_name,
                        "environment": self.environment,
                        "status": status,
                        "start_time": self.start_time_str,
                        "end_time": (self.now + datetime.timedelta(seconds=duration)).strftime("%Y-%m-%d %H:%M:%S"),
                        "duration": duration,
                        "error_reason": error_reason,
                        "error_msg": log_text,
                        "browser": self.browser,
                        "screenshot": "N/A",
                        "test_data": "Standard Automated Dataset",
                        "remarks": "Pytest Execution Verified"
                    }

                    self.test_cases.append(tc_data)

                    if module_name not in self.modules:
                        self.modules[module_name] = {"total": 0, "pass": 0, "fail": 0, "skip": 0}
                    self.modules[module_name]["total"] += 1
                    if status == "PASS":
                        self.modules[module_name]["pass"] += 1
                        self.passed_tests += 1
                    elif status == "FAIL":
                        self.modules[module_name]["fail"] += 1
                        self.failed_tests += 1
                    else:
                        self.modules[module_name]["skip"] += 1
                        self.skipped_tests += 1

                    self.total_tests += 1
                    tc_index += 1

            self._write_synced_junit_xml()
            return True
        except Exception as e:
            print(f"[ERROR] Failed parsing pytest-html: {e}")
            return False

    def _parse_junit_xml(self, xml_path):
        """Parse standard JUnit XML."""
        try:
            tree = ET.parse(xml_path)
            root = tree.getroot()
            testsuites = root.findall(".//testsuite") if root.tag != "testsuite" else [root]
            if not testsuites:
                testsuites = [root]

            tc_index = 1
            for ts in testsuites:
                for tc in ts.findall("testcase"):
                    name = tc.attrib.get("name", f"Test_{tc_index}")
                    classname = tc.attrib.get("classname", "")
                    duration = float(tc.attrib.get("time", "0.0"))
                    self.total_duration += duration

                    status = "PASS"
                    failure_reason = ""
                    error_msg = ""

                    failure = tc.find("failure")
                    error = tc.find("error")
                    skipped = tc.find("skipped")

                    if failure is not None:
                        status = "FAIL"
                        failure_reason = failure.attrib.get("message", "Assertion error")
                        error_msg = failure.text or failure_reason
                    elif error is not None:
                        status = "FAIL"
                        failure_reason = error.attrib.get("message", "Execution error")
                        error_msg = error.text or failure_reason
                    elif skipped is not None:
                        status = "SKIPPED"
                        failure_reason = skipped.attrib.get("message", "Test skipped")
                        error_msg = skipped.text or failure_reason

                    module_name = self._clean_module_name(classname or name)
                    scenario_name = self._clean_scenario_name(name)

                    tc_data = {
                        "id": f"TC_{tc_index:03d}",
                        "name": name,
                        "module": module_name,
                        "scenario": scenario_name,
                        "environment": self.environment,
                        "status": status,
                        "start_time": self.start_time_str,
                        "end_time": (self.now + datetime.timedelta(seconds=duration)).strftime("%Y-%m-%d %H:%M:%S"),
                        "duration": round(duration, 2),
                        "error_reason": failure_reason,
                        "error_msg": error_msg,
                        "browser": self.browser,
                        "screenshot": "N/A",
                        "test_data": "Standard Automated Dataset",
                        "remarks": "Pytest Execution Verified"
                    }

                    self.test_cases.append(tc_data)

                    if module_name not in self.modules:
                        self.modules[module_name] = {"total": 0, "pass": 0, "fail": 0, "skip": 0}
                    self.modules[module_name]["total"] += 1
                    if status == "PASS":
                        self.modules[module_name]["pass"] += 1
                        self.passed_tests += 1
                    elif status == "FAIL":
                        self.modules[module_name]["fail"] += 1
                        self.failed_tests += 1
                    else:
                        self.modules[module_name]["skip"] += 1
                        self.skipped_tests += 1

                    self.total_tests += 1
                    tc_index += 1
            return True
        except Exception as e:
            print(f"[ERROR] Failed parsing JUnit XML: {e}")
            return False

    def _write_synced_junit_xml(self):
        """Generate/Update results.xml from parsed test cases."""
        try:
            results_xml_path = os.path.join(self.output_dir, "results.xml")
            root = ET.Element("testsuites", name="pytest tests")
            ts = ET.SubElement(
                root, "testsuite",
                name="pytest",
                errors="0",
                failures=str(self.failed_tests),
                skipped=str(self.skipped_tests),
                tests=str(self.total_tests),
                time=str(round(self.total_duration, 3)),
                timestamp=self.start_time_str
            )
            for tc in self.test_cases:
                tc_el = ET.SubElement(
                    ts, "testcase",
                    classname=tc["module"],
                    name=tc["name"],
                    time=str(tc["duration"])
                )
                if tc["status"] == "FAIL":
                    fail_el = ET.SubElement(tc_el, "failure", message=tc["error_reason"])
                    fail_el.text = tc["error_msg"]
                elif tc["status"] == "SKIPPED":
                    skip_el = ET.SubElement(tc_el, "skipped", message=tc["error_reason"])
                    skip_el.text = tc["error_msg"]

            tree = ET.ElementTree(root)
            tree.write(results_xml_path, encoding="utf-8", xml_declaration=True)
        except Exception as e:
            print(f"[WARNING] Could not sync results.xml: {e}")

    # =========================================================================
    # 1. EXCEL REPORT GENERATION
    # =========================================================================
    def generate_excel_report(self, filename=None):
        """Generate a production-ready, beautifully styled 4-sheet Excel report."""
        safe_proj = re.sub(r'[\W_]+', '_', self.project_name)
        safe_env = re.sub(r'[\W_]+', '_', self.environment)
        if not filename:
            filename = f"{safe_proj}_{safe_env}_Automation_Report_Build_{self.build_number}.xlsx"

        excel_path = os.path.join(self.output_dir, filename)
        canonical_path = os.path.join(self.output_dir, "Test-Execution-Report.xlsx")

        wb = openpyxl.Workbook()

        # Color Palette
        navy_dark = "1A365D"
        navy_light = "2B6CB0"
        header_fill = PatternFill(start_color=navy_dark, end_color=navy_dark, fill_type="solid")
        sub_header_fill = PatternFill(start_color=navy_light, end_color=navy_light, fill_type="solid")
        kpi_fill = PatternFill(start_color="EDF2F7", end_color="EDF2F7", fill_type="solid")
        
        pass_fill = PatternFill(start_color="C6F6D5", end_color="C6F6D5", fill_type="solid")
        fail_fill = PatternFill(start_color="FED7D7", end_color="FED7D7", fill_type="solid")
        skip_fill = PatternFill(start_color="FEFCBF", end_color="FEFCBF", fill_type="solid")
        
        pass_font = Font(name="Calibri", size=10, bold=True, color="22543D")
        fail_font = Font(name="Calibri", size=10, bold=True, color="742A2A")
        skip_font = Font(name="Calibri", size=10, bold=True, color="744210")

        header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
        title_font = Font(name="Calibri", size=14, bold=True, color="FFFFFF")
        bold_font = Font(name="Calibri", size=10, bold=True, color="1A202C")
        regular_font = Font(name="Calibri", size=10, color="2D3748")

        thin_side = Side(border_style="thin", color="CBD5E0")
        border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)

        # -------------------------------------------------------------
        # SHEET 1: Execution Summary
        # -------------------------------------------------------------
        ws1 = wb.active
        ws1.title = "Execution Summary"
        ws1.views.sheetView[0].showGridLines = True

        ws1.merge_cells("A1:D1")
        ws1["A1"] = f"{self.project_name} - Test Execution Summary"
        ws1["A1"].font = title_font
        ws1["A1"].fill = header_fill
        ws1["A1"].alignment = Alignment(horizontal="center", vertical="center")
        ws1.row_dimensions[1].height = 35

        pass_pct = round((self.passed_tests / self.total_tests * 100), 2) if self.total_tests > 0 else 0.0
        fail_pct = round((self.failed_tests / self.total_tests * 100), 2) if self.total_tests > 0 else 0.0

        summary_rows = [
            ("Project Name", self.project_name, "Total Test Cases", self.total_tests),
            ("Environment", self.environment, "Passed Tests", self.passed_tests),
            ("Build Number", self.build_number, "Failed Tests", self.failed_tests),
            ("Execution Date", self.execution_date, "Skipped Tests", self.skipped_tests),
            ("Start Time", self.start_time_str, "Pass Percentage", f"{pass_pct}%"),
            ("End Time", self.end_time_str, "Fail Percentage", f"{fail_pct}%"),
            ("Total Duration", f"{round(self.total_duration, 2)} seconds", "Overall Status", self.overall_status),
            ("Jenkins Build URL", self.build_url or "N/A", "Browser", self.browser)
        ]

        ws1.append([])
        for r_idx, row in enumerate(summary_rows, start=3):
            ws1.cell(row=r_idx, column=1, value=row[0]).font = bold_font
            ws1.cell(row=r_idx, column=1).fill = kpi_fill
            ws1.cell(row=r_idx, column=2, value=row[1]).font = regular_font

            ws1.cell(row=r_idx, column=3, value=row[2]).font = bold_font
            ws1.cell(row=r_idx, column=3).fill = kpi_fill
            val_cell = ws1.cell(row=r_idx, column=4, value=row[3])
            val_cell.font = bold_font

            if row[2] == "Passed Tests": val_cell.font = pass_font; val_cell.fill = pass_fill
            elif row[2] == "Failed Tests" and self.failed_tests > 0: val_cell.font = fail_font; val_cell.fill = fail_fill
            elif row[2] == "Overall Status":
                val_cell.font = pass_font if self.overall_status == "PASSED" else fail_font
                val_cell.fill = pass_fill if self.overall_status == "PASSED" else fail_fill

            for col in range(1, 5):
                ws1.cell(row=r_idx, column=col).border = border
            ws1.row_dimensions[r_idx].height = 22

        # -------------------------------------------------------------
        # SHEET 2: Test Case Details
        # -------------------------------------------------------------
        ws2 = wb.create_sheet("Test Case Details")
        ws2.views.sheetView[0].showGridLines = True

        headers_s2 = [
            "Test Case ID", "Test Case Name", "Module", "Test Scenario", "Environment",
            "Status", "Start Time", "End Time", "Duration (s)", "Error/Failure Message",
            "Browser", "Screenshot Path", "Test Data", "Remarks"
        ]
        ws2.append(headers_s2)
        for col_num in range(1, len(headers_s2) + 1):
            cell = ws2.cell(row=1, column=col_num)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = border
        ws2.row_dimensions[1].height = 28
        ws2.freeze_panes = "A2"

        for r_idx, tc in enumerate(self.test_cases, start=2):
            row_data = [
                tc["id"], tc["name"], tc["module"], tc["scenario"], tc["environment"],
                tc["status"], tc["start_time"], tc["end_time"], tc["duration"],
                tc["error_reason"] or tc["error_msg"][:150], tc["browser"],
                tc["screenshot"] or "N/A", tc["test_data"], tc["remarks"]
            ]
            ws2.append(row_data)
            ws2.row_dimensions[r_idx].height = 20

            for col_idx in range(1, len(row_data) + 1):
                c = ws2.cell(row=r_idx, column=col_idx)
                c.font = regular_font
                c.border = border
                if col_idx == 6:
                    if tc["status"] == "PASS": c.font = pass_font; c.fill = pass_fill; c.alignment = Alignment(horizontal="center")
                    elif tc["status"] == "FAIL": c.font = fail_font; c.fill = fail_fill; c.alignment = Alignment(horizontal="center")
                    else: c.font = skip_font; c.fill = skip_fill; c.alignment = Alignment(horizontal="center")
                elif col_idx in [1, 5, 9, 11]:
                    c.alignment = Alignment(horizontal="center")

        ws2.auto_filter.ref = f"A1:{get_column_letter(len(headers_s2))}{len(self.test_cases) + 1}"

        # -------------------------------------------------------------
        # SHEET 3: Module Summary
        # -------------------------------------------------------------
        ws3 = wb.create_sheet("Module Summary")
        ws3.views.sheetView[0].showGridLines = True

        headers_s3 = ["Module Name", "Total Tests", "Passed", "Failed", "Skipped", "Pass Percentage", "Fail Percentage"]
        ws3.append(headers_s3)
        for col_num in range(1, len(headers_s3) + 1):
            cell = ws3.cell(row=1, column=col_num)
            cell.font = header_font
            cell.fill = sub_header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = border
        ws3.row_dimensions[1].height = 28
        ws3.freeze_panes = "A2"

        for r_idx, (mod_name, data) in enumerate(self.modules.items(), start=2):
            tot = data["total"]
            m_pass = data["pass"]
            m_fail = data["fail"]
            m_skip = data["skip"]
            m_pass_pct = round((m_pass / tot * 100), 2) if tot > 0 else 0.0
            m_fail_pct = round((m_fail / tot * 100), 2) if tot > 0 else 0.0

            row_data = [mod_name, tot, m_pass, m_fail, m_skip, f"{m_pass_pct}%", f"{m_fail_pct}%"]
            ws3.append(row_data)
            ws3.row_dimensions[r_idx].height = 20

            for col_idx in range(1, len(row_data) + 1):
                c = ws3.cell(row=r_idx, column=col_idx)
                c.font = regular_font
                c.border = border
                if col_idx in [2, 3, 4, 5, 6, 7]:
                    c.alignment = Alignment(horizontal="center")

        # -------------------------------------------------------------
        # SHEET 4: Failed Test Cases
        # -------------------------------------------------------------
        ws4 = wb.create_sheet("Failed Test Cases")
        ws4.views.sheetView[0].showGridLines = True

        headers_s4 = ["Test Case ID", "Test Case Name", "Module", "Failure Reason", "Error Message", "Screenshot", "Execution Time (s)", "Recommended Action"]
        ws4.append(headers_s4)
        for col_num in range(1, len(headers_s4) + 1):
            cell = ws4.cell(row=1, column=col_num)
            cell.font = header_font
            cell.fill = PatternFill(start_color="9B2C2C", end_color="9B2C2C", fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = border
        ws4.row_dimensions[1].height = 28
        ws4.freeze_panes = "A2"

        failed_list = [tc for tc in self.test_cases if tc["status"] == "FAIL"]
        if failed_list:
            for r_idx, tc in enumerate(failed_list, start=2):
                rec_action = "Inspect element locators and network response"
                if "timeout" in tc["error_reason"].lower(): rec_action = "Increase explicit wait or check page load performance"
                elif "assert" in tc["error_reason"].lower(): rec_action = "Verify expected business logic vs application state"
                elif "element" in tc["error_reason"].lower(): rec_action = "Review dynamic OutSystems DOM class or reactive rendering"

                row_data = [
                    tc["id"], tc["name"], tc["module"], tc["error_reason"],
                    tc["error_msg"][:300], tc["screenshot"] or "N/A", tc["duration"], rec_action
                ]
                ws4.append(row_data)
                ws4.row_dimensions[r_idx].height = 22
                for col_idx in range(1, len(row_data) + 1):
                    c = ws4.cell(row=r_idx, column=col_idx)
                    c.font = regular_font
                    c.border = border
                    if col_idx in [1, 7]: c.alignment = Alignment(horizontal="center")
        else:
            ws4.append(["N/A", "None - All tests passed successfully!", "N/A", "N/A", "N/A", "N/A", "N/A", "No action required"])
            for col_idx in range(1, 9):
                c = ws4.cell(row=2, column=col_idx)
                c.font = pass_font
                c.fill = pass_fill
                c.border = border

        for sheet in wb.worksheets:
            for col in sheet.columns:
                max_len = 0
                col_letter = get_column_letter(col[0].column)
                for cell in col:
                    if cell.value:
                        lines = str(cell.value).split("\n")
                        max_len = max(max_len, max(len(l) for l in lines))
                sheet.column_dimensions[col_letter].width = min(max(max_len + 4, 12), 50)

        try:
            wb.save(excel_path)
            print(f"[SUCCESS] Excel Report created at: {excel_path}")
        except Exception as e:
            print(f"[WARNING] Could not save Excel to {excel_path} (file might be open): {e}")

        try:
            wb.save(canonical_path)
            print(f"[SUCCESS] Canonical Excel Report updated at: {canonical_path}")
        except Exception as e:
            print(f"[WARNING] Could not overwrite {canonical_path} (file is currently locked or open in Excel): {e}")

        return excel_path

    # =========================================================================
    # 2. HTML REPORT GENERATION
    # =========================================================================
    def generate_html_report(self, filename=None):
        """Generate a modern, responsive, standalone HTML report with embedded SVG charts."""
        safe_proj = re.sub(r'[\W_]+', '_', self.project_name)
        safe_env = re.sub(r'[\W_]+', '_', self.environment)
        if not filename:
            filename = f"{safe_proj}_{safe_env}_Automation_Report_Build_{self.build_number}.html"

        html_path = os.path.join(self.output_dir, filename)
        canonical_path = os.path.join(self.output_dir, "report.html")

        pass_pct = round((self.passed_tests / self.total_tests * 100), 1) if self.total_tests > 0 else 0.0
        fail_pct = round((self.failed_tests / self.total_tests * 100), 1) if self.total_tests > 0 else 0.0
        skip_pct = round((self.skipped_tests / self.total_tests * 100), 1) if self.total_tests > 0 else 0.0

        circumference = 377
        pass_dash = round((pass_pct / 100) * circumference, 1)
        fail_dash = round((fail_pct / 100) * circumference, 1)
        skip_dash = round((skip_pct / 100) * circumference, 1)

        pass_offset = 0
        fail_offset = -pass_dash
        skip_offset = -(pass_dash + fail_dash)

        module_rows_html = ""
        for mod, d in self.modules.items():
            m_tot = d["total"]
            m_pass = d["pass"]
            m_fail = d["fail"]
            m_skip = d["skip"]
            m_pct = round((m_pass / m_tot * 100), 1) if m_tot > 0 else 0
            badge_color = "#38a169" if m_pct == 100 else ("#e53e3e" if m_fail > 0 else "#ecc94b")

            module_rows_html += f"""
            <tr>
                <td><strong>{mod}</strong></td>
                <td style="text-align:center;">{m_tot}</td>
                <td style="text-align:center; color:#38a169; font-weight:bold;">{m_pass}</td>
                <td style="text-align:center; color:#e53e3e; font-weight:bold;">{m_fail}</td>
                <td style="text-align:center; color:#d69e2e; font-weight:bold;">{m_skip}</td>
                <td>
                    <div style="background:#edf2f7; border-radius:999px; height:10px; width:100%; overflow:hidden;">
                        <div style="background:{badge_color}; height:100%; width:{m_pct}%;"></div>
                    </div>
                </td>
                <td style="text-align:center; font-weight:bold;">{m_pct}%</td>
            </tr>
            """

        failed_tests = [tc for tc in self.test_cases if tc["status"] == "FAIL"]
        failed_cards_html = ""
        if failed_tests:
            for tc in failed_tests:
                failed_cards_html += f"""
                <div class="fail-card">
                    <div class="fail-header">
                        <span class="badge badge-fail">FAILED</span>
                        <strong>{tc['id']}: {tc['name']}</strong>
                        <span style="float:right; font-size:12px; opacity:0.8;">Module: {tc['module']} | Duration: {tc['duration']}s</span>
                    </div>
                    <div class="fail-body">
                        <p style="margin:0 0 8px 0;"><strong>Reason:</strong> {tc['error_reason']}</p>
                        <pre>{tc['error_msg']}</pre>
                    </div>
                </div>
                """
        else:
            failed_cards_html = """
            <div style="background:#f0fff4; border:1px solid #c6f6d5; padding:18px; border-radius:8px; text-align:center; color:#22543d; font-weight:bold;">
                🎉 All test cases passed successfully with zero failures!
            </div>
            """

        test_rows_html = ""
        for tc in self.test_cases:
            badge_class = f"badge-{tc['status'].lower()}"
            error_preview = f"<div class='error-snippet'>{tc['error_reason']}</div>" if tc['error_reason'] else "-"

            test_rows_html += f"""
            <tr class="tc-row" data-status="{tc['status']}" data-module="{tc['module']}">
                <td><code>{tc['id']}</code></td>
                <td><strong>{tc['name']}</strong><br><small style="color:#718096;">{tc['scenario']}</small></td>
                <td><span class="module-tag">{tc['module']}</span></td>
                <td><span class="badge {badge_class}">{tc['status']}</span></td>
                <td>{tc['duration']}s</td>
                <td>{error_preview}</td>
                <td>{tc['browser']}</td>
            </tr>
            """

        html_content = f"""<!DOCTYPE html>
<html lang="en">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>{self.project_name} - Automation Execution Report</title>
    <style>
        :root {{
            --bg-color: #f7fafc;
            --text-color: #2d3748;
            --card-bg: #ffffff;
            --border-color: #e2e8f0;
            --primary: #1a365d;
            --primary-accent: #3182ce;
            --success: #38a169;
            --danger: #e53e3e;
            --warning: #d69e2e;
        }}
        body {{
            font-family: -apple-system, BlinkMacSystemFont, 'Segoe UI', Roboto, Helvetica, Arial, sans-serif;
            background-color: var(--bg-color);
            color: var(--text-color);
            margin: 0;
            padding: 24px;
        }}
        .container {{
            max-width: 1200px;
            margin: 0 auto;
        }}
        .header {{
            background: linear-gradient(135deg, #1a365d 0%, #2b6cb0 100%);
            color: #ffffff;
            padding: 28px 32px;
            border-radius: 12px;
            margin-bottom: 24px;
            box-shadow: 0 4px 12px rgba(0,0,0,0.08);
            display: flex;
            justify-content: space-between;
            align-items: center;
            flex-wrap: wrap;
            gap: 16px;
        }}
        .header h1 {{
            margin: 0;
            font-size: 24px;
            font-weight: 700;
        }}
        .header p {{
            margin: 6px 0 0 0;
            font-size: 14px;
            opacity: 0.9;
        }}
        .header-meta {{
            display: flex;
            gap: 12px;
        }}
        .meta-pill {{
            background: rgba(255,255,255,0.2);
            padding: 6px 14px;
            border-radius: 999px;
            font-size: 13px;
            font-weight: 600;
        }}
        .dashboard-grid {{
            display: grid;
            grid-template-columns: repeat(auto-fit, minmax(170px, 1fr));
            gap: 16px;
            margin-bottom: 24px;
        }}
        .card {{
            background: var(--card-bg);
            border-radius: 10px;
            padding: 20px;
            border: 1px solid var(--border-color);
            box-shadow: 0 2px 6px rgba(0,0,0,0.02);
            text-align: center;
        }}
        .card-label {{
            font-size: 13px;
            font-weight: 600;
            color: #718096;
            text-transform: uppercase;
            letter-spacing: 0.5px;
        }}
        .card-val {{
            font-size: 30px;
            font-weight: 800;
            margin-top: 8px;
            color: var(--primary);
        }}
        .card-val.pass {{ color: var(--success); }}
        .card-val.fail {{ color: var(--danger); }}
        .card-val.skip {{ color: var(--warning); }}

        .chart-section {{
            display: grid;
            grid-template-columns: 1fr 2fr;
            gap: 20px;
            margin-bottom: 24px;
        }}
        @media (max-width: 800px) {{
            .chart-section {{ grid-template-columns: 1fr; }}
        }}
        .chart-card {{
            background: var(--card-bg);
            border: 1px solid var(--border-color);
            border-radius: 10px;
            padding: 24px;
            box-shadow: 0 2px 6px rgba(0,0,0,0.02);
            display: flex;
            flex-direction: column;
            align-items: center;
            justify-content: center;
        }}
        .section-title {{
            font-size: 18px;
            font-weight: 700;
            margin: 0 0 16px 0;
            color: var(--primary);
            width: 100%;
        }}
        .table-responsive {{
            width: 100%;
            overflow-x: auto;
            background: var(--card-bg);
            border: 1px solid var(--border-color);
            border-radius: 10px;
            margin-bottom: 24px;
            box-shadow: 0 2px 6px rgba(0,0,0,0.02);
        }}
        table {{
            width: 100%;
            border-collapse: collapse;
            font-size: 14px;
        }}
        th, td {{
            padding: 12px 16px;
            border-bottom: 1px solid var(--border-color);
            text-align: left;
        }}
        th {{
            background-color: #edf2f7;
            font-weight: 700;
            color: #4a5568;
        }}
        tr:hover {{
            background-color: #f8fafc;
        }}
        .badge {{
            display: inline-block;
            padding: 4px 10px;
            border-radius: 6px;
            font-size: 12px;
            font-weight: 700;
        }}
        .badge-pass {{ background: #c6f6d5; color: #22543d; }}
        .badge-fail {{ background: #fed7d7; color: #742a2a; }}
        .badge-skip {{ background: #fefcbf; color: #744210; }}
        .module-tag {{
            background: #e2e8f0;
            padding: 3px 8px;
            border-radius: 4px;
            font-size: 12px;
            color: #4a5568;
        }}
        .filter-bar {{
            display: flex;
            gap: 12px;
            margin-bottom: 16px;
            flex-wrap: wrap;
            align-items: center;
        }}
        .search-box {{
            flex: 1;
            min-width: 240px;
            padding: 10px 14px;
            border-radius: 8px;
            border: 1px solid var(--border-color);
            font-size: 14px;
        }}
        .filter-btn {{
            padding: 8px 16px;
            border-radius: 6px;
            border: 1px solid var(--border-color);
            background: #ffffff;
            cursor: pointer;
            font-size: 13px;
            font-weight: 600;
        }}
        .filter-btn.active {{
            background: var(--primary);
            color: #ffffff;
            border-color: var(--primary);
        }}
        .fail-card {{
            background: #fff5f5;
            border: 1px solid #feb2b2;
            border-radius: 8px;
            margin-bottom: 12px;
            overflow: hidden;
        }}
        .fail-header {{
            background: #fed7d7;
            padding: 12px 16px;
            font-size: 14px;
            color: #742a2a;
        }}
        .fail-body {{
            padding: 16px;
            font-size: 13px;
        }}
        pre {{
            background: #2d3748;
            color: #f7fafc;
            padding: 12px;
            border-radius: 6px;
            overflow-x: auto;
            font-size: 12px;
            margin: 8px 0 0 0;
        }}
        .footer {{
            text-align: center;
            font-size: 13px;
            color: #a0aec0;
            margin-top: 32px;
        }}
    </style>
</head>
<body>
    <div class="container">
        <div class="header">
            <div>
                <h1>{self.project_name}</h1>
                <p>Automated Execution Report &bull; Environment: {self.environment} &bull; Build #{self.build_number}</p>
            </div>
            <div class="header-meta">
                <div class="meta-pill">{self.execution_date}</div>
                <div class="meta-pill" style="background:{'#38a169' if self.overall_status == 'PASSED' else '#e53e3e'};">{self.overall_status}</div>
            </div>
        </div>

        <div class="dashboard-grid">
            <div class="card">
                <div class="card-label">Total Tests</div>
                <div class="card-val">{self.total_tests}</div>
            </div>
            <div class="card">
                <div class="card-label">Passed</div>
                <div class="card-val pass">{self.passed_tests}</div>
            </div>
            <div class="card">
                <div class="card-label">Failed</div>
                <div class="card-val fail">{self.failed_tests}</div>
            </div>
            <div class="card">
                <div class="card-label">Skipped</div>
                <div class="card-val skip">{self.skipped_tests}</div>
            </div>
            <div class="card">
                <div class="card-label">Pass Rate</div>
                <div class="card-val pass">{pass_pct}%</div>
            </div>
            <div class="card">
                <div class="card-label">Total Duration</div>
                <div class="card-val">{round(self.total_duration, 1)}s</div>
            </div>
        </div>

        <div class="chart-section">
            <div class="chart-card">
                <div class="section-title">Status Breakdown</div>
                <svg width="180" height="180" viewBox="0 0 160 160">
                    <circle cx="80" cy="80" r="60" fill="transparent" stroke="#e2e8f0" stroke-width="22"/>
                    <circle cx="80" cy="80" r="60" fill="transparent" stroke="#38a169" stroke-width="22"
                            stroke-dasharray="{pass_dash} {circumference}" stroke-dashoffset="{pass_offset}" transform="rotate(-90 80 80)"/>
                    <circle cx="80" cy="80" r="60" fill="transparent" stroke="#e53e3e" stroke-width="22"
                            stroke-dasharray="{fail_dash} {circumference}" stroke-dashoffset="{fail_offset}" transform="rotate(-90 80 80)"/>
                    <circle cx="80" cy="80" r="60" fill="transparent" stroke="#d69e2e" stroke-width="22"
                            stroke-dasharray="{skip_dash} {circumference}" stroke-dashoffset="{skip_offset}" transform="rotate(-90 80 80)"/>
                    <text x="80" y="86" text-anchor="middle" font-size="20" font-weight="bold" fill="#1a365d">{pass_pct}%</text>
                </svg>
            </div>
            <div class="chart-card" style="align-items: stretch;">
                <div class="section-title">Module Summary</div>
                <div class="table-responsive" style="margin: 0; border: none; box-shadow: none;">
                    <table>
                        <thead>
                            <tr>
                                <th>Module</th>
                                <th style="text-align:center;">Total</th>
                                <th style="text-align:center;">Pass</th>
                                <th style="text-align:center;">Fail</th>
                                <th style="text-align:center;">Skip</th>
                                <th style="width: 140px;">Progress</th>
                                <th style="text-align:center;">Pass %</th>
                            </tr>
                        </thead>
                        <tbody>
                            {module_rows_html}
                        </tbody>
                    </table>
                </div>
            </div>
        </div>

        <div class="section-title" style="margin-top: 24px;">Failed Test Analysis</div>
        <div style="margin-bottom: 24px;">
            {failed_cards_html}
        </div>

        <div class="section-title">Detailed Test Execution</div>
        <div class="filter-bar">
            <input type="text" id="searchInput" class="search-box" placeholder="Search test name, scenario, or module..." onkeyup="filterTests()">
            <button class="filter-btn active" onclick="setStatusFilter('ALL', this)">All ({self.total_tests})</button>
            <button class="filter-btn" onclick="setStatusFilter('PASS', this)">Passed ({self.passed_tests})</button>
            <button class="filter-btn" onclick="setStatusFilter('FAIL', this)">Failed ({self.failed_tests})</button>
            <button class="filter-btn" onclick="setStatusFilter('SKIPPED', this)">Skipped ({self.skipped_tests})</button>
        </div>

        <div class="table-responsive">
            <table id="testTable">
                <thead>
                    <tr>
                        <th>ID</th>
                        <th>Test Case & Scenario</th>
                        <th>Module</th>
                        <th>Status</th>
                        <th>Duration</th>
                        <th>Failure Summary</th>
                        <th>Browser</th>
                    </tr>
                </thead>
                <tbody>
                    {test_rows_html}
                </tbody>
            </table>
        </div>

        <div class="footer">
            Generated automatically by Legalhub CI/CD Reporting Engine &bull; {self.start_time_str}
        </div>
    </div>

    <script>
        let currentStatusFilter = 'ALL';

        function setStatusFilter(status, btn) {{
            currentStatusFilter = status;
            document.querySelectorAll('.filter-btn').forEach(b => b.classList.remove('active'));
            btn.classList.add('active');
            filterTests();
        }}

        function filterTests() {{
            const search = document.getElementById('searchInput').value.toLowerCase();
            const rows = document.querySelectorAll('.tc-row');

            rows.forEach(row => {{
                const status = row.getAttribute('data-status');
                const text = row.innerText.toLowerCase();

                const matchesStatus = (currentStatusFilter === 'ALL' || status === currentStatusFilter);
                const matchesSearch = text.includes(search);

                if (matchesStatus && matchesSearch) {{
                    row.style.display = '';
                }} else {{
                    row.style.display = 'none';
                }}
            }});
        }}
    </script>
</body>
</html>
"""
        try:
            with open(html_path, "w", encoding="utf-8") as f:
                f.write(html_content)
            print(f"[SUCCESS] HTML Report created at: {html_path}")
        except Exception as e:
            print(f"[WARNING] Could not write HTML report to {html_path}: {e}")

        try:
            with open(canonical_path, "w", encoding="utf-8") as f:
                f.write(html_content)
            print(f"[SUCCESS] Canonical HTML Report updated at: {canonical_path}")
        except Exception as e:
            print(f"[WARNING] Could not write HTML report to {canonical_path}: {e}")

        return html_path

    # =========================================================================
    # 3. EMAIL BODY HTML GENERATION
    # =========================================================================
    def generate_email_body_html(self):
        """Generate email-client-compatible HTML matching the exact user specification."""
        pass_pct = round((self.passed_tests / self.total_tests * 100), 2) if self.total_tests > 0 else 0.0
        
        # Duration formatting (e.g. 28m 30s)
        mins = int(self.total_duration // 60)
        secs = int(self.total_duration % 60)
        dur_formatted = f"{mins}m {secs}s" if mins > 0 else f"{secs}s"

        status_color = "#28a745" if self.overall_status == "PASSED" else "#dc3545"

        failed_list = [tc for tc in self.test_cases if tc["status"] == "FAIL"]
        if failed_list:
            failed_rows = ""
            for tc in failed_list:
                failed_rows += f"""
                <tr>
                    <td style="padding: 8px 12px; border: 1px solid #e2e8f0; font-size: 13px;"><code>{tc['id']}</code></td>
                    <td style="padding: 8px 12px; border: 1px solid #e2e8f0; font-size: 13px; font-weight: bold;">{tc['name']}</td>
                    <td style="padding: 8px 12px; border: 1px solid #e2e8f0; font-size: 13px;">{tc['module']}</td>
                    <td style="padding: 8px 12px; border: 1px solid #e2e8f0; font-size: 13px; color: #dc3545;">{tc['error_reason']}</td>
                    <td style="padding: 8px 12px; border: 1px solid #e2e8f0; font-size: 12px; font-family: monospace; background-color: #f8f9fa;">{tc['error_msg'][:250]}</td>
                </tr>
                """
            failure_section_html = f"""
            <h3 style="color: #742a2a; margin-top: 25px; margin-bottom: 10px; font-size: 16px; border-bottom: 2px solid #feb2b2; padding-bottom: 6px;">Failed Test Cases</h3>
            <table style="width: 100%; border-collapse: collapse; margin-bottom: 20px; font-family: 'Segoe UI', Arial, sans-serif;">
                <thead>
                    <tr style="background-color: #fed7d7; color: #742a2a;">
                        <th style="padding: 10px; border: 1px solid #feb2b2; text-align: left; font-size: 13px;">Test Case ID</th>
                        <th style="padding: 10px; border: 1px solid #feb2b2; text-align: left; font-size: 13px;">Test Case Name</th>
                        <th style="padding: 10px; border: 1px solid #feb2b2; text-align: left; font-size: 13px;">Module</th>
                        <th style="padding: 10px; border: 1px solid #feb2b2; text-align: left; font-size: 13px;">Failure Reason</th>
                        <th style="padding: 10px; border: 1px solid #feb2b2; text-align: left; font-size: 13px;">Error Message</th>
                    </tr>
                </thead>
                <tbody>
                    {failed_rows}
                </tbody>
            </table>
            """
        else:
            failure_section_html = """
            <h3 style="color: #2d3748; margin-top: 25px; margin-bottom: 10px; font-size: 16px; border-bottom: 2px solid #e2e8f0; padding-bottom: 6px;">Failed Test Cases</h3>
            <p style="margin: 12px 0 20px 0; color: #28a745; font-weight: bold; background-color: #f0fff4; border: 1px solid #c6f6d5; padding: 12px 16px; border-radius: 6px;">
                No test cases failed during this execution.
            </p>
            """

        email_html = f"""<!DOCTYPE html>
<html>
<head>
    <meta charset="utf-8">
</head>
<body style="font-family: 'Segoe UI', Arial, Helvetica, sans-serif; background-color: #f4f6f9; margin: 0; padding: 20px; color: #333333; line-height: 1.5;">
    <table align="center" border="0" cellpadding="0" cellspacing="0" width="680" style="background-color: #ffffff; border-radius: 8px; overflow: hidden; box-shadow: 0 4px 12px rgba(0,0,0,0.08); border: 1px solid #e1e4e8;">
        <!-- Header -->
        <tr>
            <td style="background-color: #1a365d; padding: 24px 30px; color: #ffffff; text-align: center;">
                <h1 style="margin: 0; font-size: 22px; font-weight: 700; letter-spacing: 0.5px;">[Automation Test Report] {self.project_name}</h1>
                <p style="margin: 6px 0 0 0; font-size: 14px; opacity: 0.9;">Environment: <strong>{self.environment}</strong> &bull; Build <strong>#{self.build_number}</strong></p>
            </td>
        </tr>
        
        <!-- Content -->
        <tr>
            <td style="padding: 28px 32px;">
                <p style="font-size: 15px; margin: 0 0 10px 0;"><strong>Hello Team,</strong></p>
                <p style="font-size: 14px; color: #4a5568; margin: 0 0 20px 0;">The automation test execution has been completed. Please find the execution summary below.</p>

                <!-- Execution Summary Section -->
                <h3 style="color: #1a365d; margin: 0 0 10px 0; font-size: 16px; border-bottom: 2px solid #e2e8f0; padding-bottom: 6px;">Execution Summary</h3>
                <table width="100%" cellpadding="8" cellspacing="0" style="border-collapse: collapse; margin-bottom: 24px; font-size: 14px;">
                    <tr style="background-color: #f7fafc;">
                        <th width="35%" style="text-align: left; padding: 10px 12px; border: 1px solid #e2e8f0; color: #4a5568;">Field</th>
                        <th width="65%" style="text-align: left; padding: 10px 12px; border: 1px solid #e2e8f0; color: #4a5568;">Details</th>
                    </tr>
                    <tr>
                        <td style="padding: 8px 12px; border: 1px solid #e2e8f0; font-weight: 600; color: #4a5568;">Project</td>
                        <td style="padding: 8px 12px; border: 1px solid #e2e8f0;">{self.project_name}</td>
                    </tr>
                    <tr style="background-color: #fcfdfe;">
                        <td style="padding: 8px 12px; border: 1px solid #e2e8f0; font-weight: 600; color: #4a5568;">Environment</td>
                        <td style="padding: 8px 12px; border: 1px solid #e2e8f0;">{self.environment}</td>
                    </tr>
                    <tr>
                        <td style="padding: 8px 12px; border: 1px solid #e2e8f0; font-weight: 600; color: #4a5568;">Build Number</td>
                        <td style="padding: 8px 12px; border: 1px solid #e2e8f0;">#{self.build_number}</td>
                    </tr>
                    <tr style="background-color: #fcfdfe;">
                        <td style="padding: 8px 12px; border: 1px solid #e2e8f0; font-weight: 600; color: #4a5568;">Execution Status</td>
                        <td style="padding: 8px 12px; border: 1px solid #e2e8f0;"><strong style="color: {status_color}; font-size: 14px;">{self.overall_status}</strong></td>
                    </tr>
                    <tr>
                        <td style="padding: 8px 12px; border: 1px solid #e2e8f0; font-weight: 600; color: #4a5568;">Execution Date</td>
                        <td style="padding: 8px 12px; border: 1px solid #e2e8f0;">{self.execution_date}</td>
                    </tr>
                    <tr style="background-color: #fcfdfe;">
                        <td style="padding: 8px 12px; border: 1px solid #e2e8f0; font-weight: 600; color: #4a5568;">Start Time</td>
                        <td style="padding: 8px 12px; border: 1px solid #e2e8f0;">{self.start_time_str}</td>
                    </tr>
                    <tr>
                        <td style="padding: 8px 12px; border: 1px solid #e2e8f0; font-weight: 600; color: #4a5568;">End Time</td>
                        <td style="padding: 8px 12px; border: 1px solid #e2e8f0;">{self.end_time_str}</td>
                    </tr>
                    <tr style="background-color: #fcfdfe;">
                        <td style="padding: 8px 12px; border: 1px solid #e2e8f0; font-weight: 600; color: #4a5568;">Total Duration</td>
                        <td style="padding: 8px 12px; border: 1px solid #e2e8f0;">{dur_formatted} ({round(self.total_duration, 1)} seconds)</td>
                    </tr>
                </table>

                <!-- Test Summary Section -->
                <h3 style="color: #1a365d; margin: 0 0 10px 0; font-size: 16px; border-bottom: 2px solid #e2e8f0; padding-bottom: 6px;">Test Summary</h3>
                <table width="100%" cellpadding="8" cellspacing="0" style="border-collapse: collapse; margin-bottom: 24px; font-size: 14px;">
                    <tr style="background-color: #f7fafc;">
                        <th width="65%" style="text-align: left; padding: 10px 12px; border: 1px solid #e2e8f0; color: #4a5568;">Metric</th>
                        <th width="35%" style="text-align: right; padding: 10px 12px; border: 1px solid #e2e8f0; color: #4a5568;">Count</th>
                    </tr>
                    <tr>
                        <td style="padding: 8px 12px; border: 1px solid #e2e8f0; font-weight: 600;">Total Test Cases</td>
                        <td style="padding: 8px 12px; border: 1px solid #e2e8f0; text-align: right; font-weight: bold;">{self.total_tests}</td>
                    </tr>
                    <tr style="background-color: #f0fff4;">
                        <td style="padding: 8px 12px; border: 1px solid #e2e8f0; font-weight: 600; color: #22543d;">Passed</td>
                        <td style="padding: 8px 12px; border: 1px solid #e2e8f0; text-align: right; font-weight: bold; color: #28a745;">{self.passed_tests}</td>
                    </tr>
                    <tr style="background-color: {'#fff5f5' if self.failed_tests > 0 else '#ffffff'};">
                        <td style="padding: 8px 12px; border: 1px solid #e2e8f0; font-weight: 600; color: {'#742a2a' if self.failed_tests > 0 else '#4a5568'};">Failed</td>
                        <td style="padding: 8px 12px; border: 1px solid #e2e8f0; text-align: right; font-weight: bold; color: {'#dc3545' if self.failed_tests > 0 else '#4a5568'};">{self.failed_tests}</td>
                    </tr>
                    <tr style="background-color: {'#fefcbf' if self.skipped_tests > 0 else '#ffffff'};">
                        <td style="padding: 8px 12px; border: 1px solid #e2e8f0; font-weight: 600; color: {'#744210' if self.skipped_tests > 0 else '#4a5568'};">Skipped</td>
                        <td style="padding: 8px 12px; border: 1px solid #e2e8f0; text-align: right; font-weight: bold; color: {'#d69e2e' if self.skipped_tests > 0 else '#4a5568'};">{self.skipped_tests}</td>
                    </tr>
                    <tr style="background-color: #edf2f7;">
                        <td style="padding: 8px 12px; border: 1px solid #e2e8f0; font-weight: 700; color: #1a365d;">Pass Percentage</td>
                        <td style="padding: 8px 12px; border: 1px solid #e2e8f0; text-align: right; font-weight: 700; color: #1a365d;">{pass_pct}%</td>
                    </tr>
                </table>

                <!-- Failure Information Section -->
                {failure_section_html}

                <!-- Jenkins Information Section -->
                <h3 style="color: #1a365d; margin: 0 0 10px 0; font-size: 16px; border-bottom: 2px solid #e2e8f0; padding-bottom: 6px;">Jenkins Build Information</h3>
                <table width="100%" cellpadding="8" cellspacing="0" style="border-collapse: collapse; margin-bottom: 20px; font-size: 14px;">
                    <tr>
                        <td width="35%" style="padding: 8px 12px; border: 1px solid #e2e8f0; font-weight: 600; color: #4a5568;">Build Number</td>
                        <td width="65%" style="padding: 8px 12px; border: 1px solid #e2e8f0;">#{self.build_number}</td>
                    </tr>
                    <tr style="background-color: #fcfdfe;">
                        <td style="padding: 8px 12px; border: 1px solid #e2e8f0; font-weight: 600; color: #4a5568;">Build Status</td>
                        <td style="padding: 8px 12px; border: 1px solid #e2e8f0;"><strong style="color: {status_color};">{self.overall_status}</strong></td>
                    </tr>
                    <tr>
                        <td style="padding: 8px 12px; border: 1px solid #e2e8f0; font-weight: 600; color: #4a5568;">Jenkins Job Name</td>
                        <td style="padding: 8px 12px; border: 1px solid #e2e8f0;">{self.job_name}</td>
                    </tr>
                    <tr style="background-color: #fcfdfe;">
                        <td style="padding: 8px 12px; border: 1px solid #e2e8f0; font-weight: 600; color: #4a5568;">Jenkins Build URL</td>
                        <td style="padding: 8px 12px; border: 1px solid #e2e8f0;"><a href="{self.build_url}" style="color: #3182ce; text-decoration: none;">{self.build_url or 'N/A'}</a></td>
                    </tr>
                </table>

                <!-- CTA Button -->
                <div style="text-align: center; margin: 25px 0 10px 0;">
                    <a href="{self.build_url}" style="background-color: #3182ce; color: #ffffff; text-decoration: none; padding: 12px 28px; border-radius: 6px; font-weight: 600; font-size: 14px; display: inline-block;">View Jenkins Build</a>
                </div>
            </td>
        </tr>

        <!-- Footer -->
        <tr>
            <td style="background-color: #fafafa; padding: 20px 32px; border-top: 1px solid #edf2f7; font-size: 13px; color: #4a5568;">
                <p style="margin: 0 0 4px 0;"><strong>Regards,</strong></p>
                <p style="margin: 0 0 12px 0; font-weight: 600; color: #1a365d;">Automation Testing Team</p>
                <p style="margin: 0; font-size: 12px; color: #a0aec0; font-style: italic;">This is an automated email generated by the Jenkins automation pipeline. Please do not reply directly to this email.</p>
            </td>
        </tr>
    </table>
</body>
</html>
"""
        email_body_path = os.path.join(self.output_dir, "email_notification.html")
        try:
            with open(email_body_path, "w", encoding="utf-8") as f:
                f.write(email_html)
            print(f"[SUCCESS] Email Notification HTML created at: {email_body_path}")
        except Exception as e:
            print(f"[WARNING] Could not write email body: {e}")
        return email_body_path

    # =========================================================================
    # 4. JSON SUMMARY GENERATION
    # =========================================================================
    def generate_summary_json(self):
        """Generate machine-readable summary JSON for Jenkins pipeline."""
        pass_pct = round((self.passed_tests / self.total_tests * 100), 2) if self.total_tests > 0 else 0.0
        safe_proj = re.sub(r'[\W_]+', '_', self.project_name)
        safe_env = re.sub(r'[\W_]+', '_', self.environment)
        
        summary_data = {
            "project_name": self.project_name,
            "environment": self.environment,
            "build_number": self.build_number,
            "execution_date": self.execution_date,
            "start_time": self.start_time_str,
            "end_time": self.end_time_str,
            "total_duration": self.total_duration,
            "total_tests": self.total_tests,
            "passed": self.passed_tests,
            "failed": self.failed_tests,
            "skipped": self.skipped_tests,
            "pass_percentage": f"{pass_pct}%",
            "overall_status": self.overall_status,
            "excel_report": f"{safe_proj}_{safe_env}_Automation_Report_Build_{self.build_number}.xlsx",
            "html_report": f"{safe_proj}_{safe_env}_Automation_Report_Build_{self.build_number}.html",
            "failed_cases": [tc for tc in self.test_cases if tc["status"] == "FAIL"]
        }

        json_path = os.path.join(self.output_dir, "summary.json")
        try:
            with open(json_path, "w", encoding="utf-8") as f:
                json.dump(summary_data, f, indent=2)
            print(f"[SUCCESS] Summary JSON created at: {json_path}")
        except Exception as e:
            print(f"[WARNING] Could not write summary JSON: {e}")
        return json_path

    def generate_all_reports(self):
        """Generate Excel, HTML, Email Body, and Summary JSON reports."""
        excel_file = self.generate_excel_report()
        html_file = self.generate_html_report()
        self.generate_email_body_html()
        self.generate_summary_json()
        return excel_file, html_file


def main():
    source_path = sys.argv[1] if len(sys.argv) > 1 else None
    output_dir = sys.argv[2] if len(sys.argv) > 2 else "reports"

    generator = ReportGenerator(source_path=source_path, output_dir=output_dir)
    excel_file, html_file = generator.generate_all_reports()
    print(f"\n==================================================")
    print(f"[REPORT GENERATION COMPLETED]")
    print(f"  - Total Tests:     {generator.total_tests}")
    print(f"  - Passed:          {generator.passed_tests}")
    print(f"  - Failed:          {generator.failed_tests}")
    print(f"  - Skipped:         {generator.skipped_tests}")
    print(f"  - Overall Status:  {generator.overall_status}")
    print(f"  - Excel Report:    {excel_file}")
    print(f"  - HTML Report:     {html_file}")
    print(f"==================================================")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
